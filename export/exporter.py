import pandas as pd
import os
from openpyxl.chart import Reference, ScatterChart, Series
from engine.levels import format_levels_text
from app_info import APP_VERSION


def _metadata_rows(result):
    return [
        ("Method", result.get("method_label", result.get("method", "-"))),
        ("Tide Type", result.get("type", "-")),
        ("Formzahl", result.get("formzahl")),
        ("RMSE (m)", result.get("rmse")),
        ("R2", result.get("r2", "-")),
        ("RMSE % Range", result.get("rmse_percent_range", "-")),
        ("Quality", result.get("quality", "-")),
        ("Source File", result.get("source_file", "-")),
        ("Source Path", result.get("source_path", "-")),
        ("Data Points", result.get("data_points", "-")),
        ("Data Duration (days)", result.get("data_days", "-")),
        ("Data Start", result.get("data_start", "-")),
        ("Data End", result.get("data_end", "-")),
        ("Developer", result.get("developer", "dhanyxyz1-atl")),
        ("App Version", result.get("app_version", APP_VERSION)),
        ("Analysis Note", result.get("analysis_note", "-")),
    ]


def export_results(result, filepath, predicted_df=None):
    """
    Mengekspor hasil analisis pasut ke file Excel (.xlsx) atau CSV (.csv).

    Args:
        result (dict): Output dari run_analysis (berisi 'constituents', 'levels', dst).
            'levels' berformat {"HWS": (value, count), ...}
        filepath (str): Path tujuan file, ekstensi menentukan format (.xlsx atau .csv).
        predicted_df (pd.DataFrame, optional): DataFrame prediksi (kolom 'time', 'height').
    """
    ext = os.path.splitext(filepath)[1].lower()

    df_const = result['constituents'].copy()

    levels = result['levels']
    df_levels = pd.DataFrame([
        {"Level": key, "Value (m)": value, "Event Count": count}
        for key, (value, count) in levels.items()
    ])

    df_summary = pd.DataFrame([{
        "Method": result.get("method_label", result.get("method", "-")),
        "Source File": result.get("source_file", "-"),
        "Data Duration (days)": result.get("data_days", "-"),
        "Developer": result.get("developer", "dhanyxyz1-atl"),
        "App Version": result.get("app_version", APP_VERSION),
        "Formzahl": result.get('formzahl'),
        "Tide Type": result.get('type'),
        "RMSE": result.get('rmse'),
        "R2": result.get("r2", "-"),
        "RMSE % Range": result.get("rmse_percent_range", "-"),
        "Quality": result.get("quality", "-"),
    }])
    df_metadata = pd.DataFrame(_metadata_rows(result), columns=["Item", "Value"])

    levels_text = format_levels_text(levels)
    levels_report_lines = [
        f"Method                 : {result.get('method_label', result.get('method', '-'))}",
        f"Source File            : {result.get('source_file', '-')}",
        f"Data Duration          : {result.get('data_days', '-')} days",
        f"Developer              : {result.get('developer', 'dhanyxyz1-atl')}",
        f"App Version            : {result.get('app_version', APP_VERSION)}",
        f"R2                     : {result.get('r2', '-')}",
        f"RMSE Percent Range     : {result.get('rmse_percent_range', '-')} %",
        f"Quality                : {result.get('quality', '-')}",
        "",
        *levels_text.split("\n"),
    ]

    if ext == '.xlsx':
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            df_metadata.to_excel(writer, sheet_name='Metadata', index=False)
            df_const.to_excel(writer, sheet_name='Constituents', index=False)
            df_levels.to_excel(writer, sheet_name='Important Levels', index=False)

            df_levels_text = pd.DataFrame({"Important Levels Report": levels_report_lines})
            df_levels_text.to_excel(writer, sheet_name='Levels Report', index=False)

            if predicted_df is not None:
                predicted_df.to_excel(writer, sheet_name='Prediction', index=False)

            observed_prediction = result.get("observed_prediction")
            if observed_prediction is not None and not observed_prediction.empty:
                observed_prediction.to_excel(writer, sheet_name="Observed vs Predicted", index=False)

            for sheet in writer.book.worksheets:
                _format_sheet(sheet)

            if "Observed vs Predicted" in writer.book.sheetnames:
                _add_xy_chart(
                    writer.book["Observed vs Predicted"],
                    title="Observed Tide vs Predicted Tide",
                    series_pairs=[(1, 2), (1, 3)],
                    anchor="F2",
                )

            if predicted_df is not None and "Prediction" in writer.book.sheetnames:
                _add_xy_chart(
                    writer.book["Prediction"],
                    title="1-Year Tide Prediction",
                    series_pairs=[(1, 2)],
                    anchor="D2",
                )

    elif ext == '.csv':
        base, _ = os.path.splitext(filepath)
        df_summary.to_csv(f"{base}_summary.csv", index=False)
        df_metadata.to_csv(f"{base}_metadata.csv", index=False)
        df_const.to_csv(f"{base}_constituents.csv", index=False)
        df_levels.to_csv(f"{base}_levels.csv", index=False)

        with open(f"{base}_levels_report.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(levels_report_lines))

        if predicted_df is not None:
            predicted_df.to_csv(f"{base}_prediction.csv", index=False)
    else:
        raise ValueError("Unsupported file format. Use .xlsx or .csv.")


def export_preprocess_result(payload, filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext != ".xlsx":
        raise ValueError("Invert & Shift export supports .xlsx only.")

    source = payload["source"]
    processed = payload["processed"]
    reference = payload.get("reference")
    source_stats = payload.get("source_stats", {})
    reference_stats = payload.get("reference_stats")

    summary_rows = [
        ("Process", "Invert & Shift Tide"),
        ("Source Points", source_stats.get("points", len(source))),
        ("Source HWL", source_stats.get("hwl")),
        ("Source LWL", source_stats.get("lwl")),
        ("Source MSL", source_stats.get("msl")),
        ("Reference Points", reference_stats.get("points") if reference_stats else "-"),
        ("Reference HWL", reference_stats.get("hwl") if reference_stats else "-"),
        ("Reference LWL", reference_stats.get("lwl") if reference_stats else "-"),
        ("Reference MSL", reference_stats.get("msl") if reference_stats else "-"),
        ("Output Points", len(processed)),
        ("Note", payload.get("note", "-")),
    ]
    summary = pd.DataFrame(summary_rows, columns=["Item", "Value"])
    chart_data = _build_preprocess_chart_data(source, processed, reference)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        processed.to_excel(writer, sheet_name="Processed Tide", index=False)
        source.to_excel(writer, sheet_name="Source Tide", index=False)
        if reference is not None and not reference.empty:
            reference.to_excel(writer, sheet_name="Reference Tide", index=False)
        chart_data.to_excel(writer, sheet_name="Chart Data", index=False)

        for sheet in writer.book.worksheets:
            _format_sheet(sheet)

        _add_xy_chart(
            writer.book["Chart Data"],
            title="Invert & Shift Tide Chart",
            series_pairs=[(1, 2), (3, 4), (5, 6), (7, 8)],
            anchor="J2",
        )


def _build_preprocess_chart_data(source, processed, reference):
    max_len = max(
        len(source),
        len(processed),
        len(reference) if reference is not None else 0,
    )

    def pad(series, length=max_len):
        return series.reset_index(drop=True).reindex(range(length))

    reference = reference if reference is not None else pd.DataFrame(columns=["time", "height"])
    return pd.DataFrame({
        "raw_time": pad(source["time"]),
        "raw_height": pad(source["height"]),
        "inverted_time": pad(processed["time"]),
        "inverted_height": pad(processed["inverted_height"]),
        "reference_time": pad(reference["time"]) if not reference.empty else pd.Series([None] * max_len),
        "reference_height": pad(reference["height"]) if not reference.empty else pd.Series([None] * max_len),
        "shifted_time": pad(processed["time"]),
        "shifted_height": pad(processed["shifted_height"]),
    })


def _format_sheet(sheet):
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)


def _add_xy_chart(sheet, title, series_pairs, anchor):
    if sheet.max_row < 3:
        return

    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.y_axis.title = "Water Level (m)"
    chart.x_axis.title = "Time"
    chart.height = 12
    chart.width = 24
    chart.scatterStyle = "lineMarker"

    for x_col, y_col in series_pairs:
        if x_col <= sheet.max_column and y_col <= sheet.max_column:
            xvalues = Reference(sheet, min_col=x_col, min_row=2, max_row=sheet.max_row)
            yvalues = Reference(sheet, min_col=y_col, min_row=2, max_row=sheet.max_row)
            title = sheet.cell(row=1, column=y_col).value
            chart.series.append(Series(yvalues, xvalues, title=title))
    sheet.add_chart(chart, anchor)
